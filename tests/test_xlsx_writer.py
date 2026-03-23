from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from backend.infrastructure.db import SQLiteAnalysisRepository, create_connection
from backend.infrastructure.xlsx_writer import XlsxReportWriter


def test_xlsx_writer_builds_zero_filled_line_counts(tmp_path):
    db_path = tmp_path / "test.db"
    conn = create_connection(db_path)
    repo = SQLiteAnalysisRepository(conn)
    repo.init()

    analysis_id = "analysis-1"
    source_path = tmp_path / "source.txt"
    source_path.write_text("demo", encoding="utf-8")

    repo.create_analysis(
        analysis_id=analysis_id,
        file_path=str(source_path),
        target_lemma="житель",
    )
    repo.upsert_word_totals_batch(analysis_id, {"житель": 3})
    repo.upsert_word_line_counts_batch(analysis_id, 2, {"житель": 2})
    repo.upsert_word_line_counts_batch(analysis_id, 4, {"житель": 1})
    repo.mark_success(analysis_id, total_lines=5)

    writer = XlsxReportWriter()
    data = writer.generate_report_xlsx_bytes(analysis_id, repo)

    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws["A2"].value == "житель"
    assert ws["B2"].value == 3
    assert ws["C2"].value == "0,2,0,1,0"
    conn.close()
